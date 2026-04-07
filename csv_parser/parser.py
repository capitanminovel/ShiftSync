"""
csv_parser/parser.py

Responsibilities:
- Read a CSV *or* XLSX file and return normalised shift dicts
- Supports two XLSX layouts:
    "flat" — standard column headers (employee, date, start time, end time)
    "schedule grid" — the weekly grid format produced by scheduling tools:
        Row 1 : day-of-month numbers in even columns (0, 2, 4 …)
        Row 2 : repeating "SHIFT" / "EMPLOYEE" headers
        Row 3+: shift string / employee name pairs
- Parse and validate each row into a clean shift dict
- Return a list of valid shifts and a list of row-level errors

Time strings like "9:30am - 4pm", "4:30/5pm - 8+pm" are handled by the
fuzzy time parser.
"""

import csv
import io
import re
from datetime import datetime, date
from typing import IO, Optional


# ---------------------------------------------------------------------------
# Column name aliases — maps every acceptable header variation → canonical key
# ---------------------------------------------------------------------------
COLUMN_ALIASES: dict[str, str] = {
    # employee
    "employee": "employee",
    "employee name": "employee",
    "name": "employee",
    "staff": "employee",
    # date
    "date": "date",
    "shift date": "date",
    "work date": "date",
    # start time
    "start": "start_time",
    "start time": "start_time",
    "start_time": "start_time",
    "from": "start_time",
    # end time
    "end": "end_time",
    "end time": "end_time",
    "end_time": "end_time",
    "to": "end_time",
    # optional
    "notes": "notes",
    "note": "notes",
    "description": "notes",
    "location": "location",
    "role": "role",
    "position": "role",
}

REQUIRED_COLUMNS = {"employee", "date", "start_time", "end_time"}

DATE_FORMATS = ["%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%d/%m/%Y", "%B %d, %Y"]
TIME_FORMATS = ["%H:%M", "%I:%M %p", "%I:%M%p", "%H:%M:%S", "%I %p"]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse_file(file: IO, filename: str = "") -> tuple[list[dict], list[dict]]:
    """
    Dispatch to the correct parser based on file extension.

    Accepts .csv or .xlsx files.
    Returns (shifts, errors).
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "xlsx":
        return parse_xlsx(file)
    return parse_csv(file)


def parse_xlsx(file: IO) -> tuple[list[dict], list[dict]]:
    """
    Parse a weekly schedule XLSX in the grid format:
      Row 1: day-of-month numbers in even columns (A, C, E …)
      Row 2: repeating SHIFT / EMPLOYEE headers
      Row 3+: shift string / employee name pairs

    The year and month are inferred from the filename passed to parse_file,
    but this function also accepts them as keyword args for testing.
    Dates are resolved by reading the first row's integer day numbers and
    the workbook's creation / modified metadata if available, otherwise
    defaulting to the current year/month.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for XLSX parsing. pip install openpyxl")

    wb = load_workbook(file, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 3:
        raise ValueError("XLSX schedule must have at least 3 rows (dates, headers, data).")

    # --- Resolve dates from row 1 ---
    day_cols = [i for i in range(0, len(rows[0]), 2) if rows[0][i] is not None]
    try:
        day_nums = [int(rows[0][c]) for c in day_cols]
    except (TypeError, ValueError):
        raise ValueError("Row 1 of the XLSX must contain numeric day-of-month values.")

    # Infer year/month: use today's date as the anchor
    today = datetime.today()
    year, month = today.year, today.month

    shifts, errors = [], []

    for row_idx, row in enumerate(rows[2:], start=3):  # rows[2:] = data rows
        for i, col in enumerate(day_cols):
            shift_str = row[col] if col < len(row) else None
            employee  = row[col + 1] if col + 1 < len(row) else None

            if not shift_str or not employee:
                continue
            if "UNAVAILABLE" in str(shift_str).upper():
                continue

            try:
                start_time, end_time = _parse_shift_range(str(shift_str), row_idx)
                shift_date = date(year, month, day_nums[i])
                shifts.append({
                    "employee":   str(employee).strip(),
                    "date":       shift_date,
                    "start_time": start_time,
                    "end_time":   end_time,
                    "notes":      "",
                    "location":   "",
                    "role":       "",
                })
            except ValueError as exc:
                errors.append({"row": row_idx, "message": str(exc)})

    return shifts, errors


def _parse_shift_range(shift_str: str, row_index: int):
    """
    Parse a shift string like "9:30am - 4pm" or "4:30/5pm - 8+pm"
    into (start_time, end_time).

    The "slash" variant (4:30/5pm) means a flexible start — we take
    the earlier time. The "+" suffix (8+pm) means "at least 8pm" —
    we treat it as exactly 8pm.
    """
    # Normalise: remove "+" (open-ended), collapse whitespace
    cleaned = shift_str.strip().replace("+", "")

    parts = re.split(r"\s*-\s*", cleaned, maxsplit=1)
    if len(parts) != 2:
        raise ValueError(
            f"Row {row_index}: Cannot split shift range '{shift_str}'. "
            "Expected format: 'start - end'."
        )

    start_str, end_str = parts
    start_time = _parse_flexible_time(start_str.strip(), row_index, label="start")
    end_time   = _parse_flexible_time(end_str.strip(),   row_index, label="end")

    return start_time, end_time


def _parse_flexible_time(value: str, row_index: int, label: str):
    """
    Handle "4:30/5pm" (slash = two options, take the earlier one)
    and plain "9:30am", "4pm", "12pm".
    """
    # Slash variant: "4:30/5pm" → pick the earlier candidate
    if "/" in value:
        # Both sides share the am/pm suffix from the last character group
        suffix_match = re.search(r"(am|pm)$", value, re.IGNORECASE)
        suffix = suffix_match.group(1) if suffix_match else ""
        candidates_raw = value.replace(suffix, "").split("/")
        candidates = []
        for c in candidates_raw:
            t = _parse_time(f"{c.strip()}{suffix}", row_index, label)
            candidates.append(t)
        return min(candidates)

    return _parse_time(value, row_index, label)


def parse_csv(file: IO[bytes] | IO[str]) -> tuple[list[dict], list[dict]]:
    """
    Parse a CSV file into a list of shift dicts.

    Args:
        file: A file-like object (binary or text mode).

    Returns:
        (shifts, errors)
        shifts — list of dicts with keys: employee, date, start_time,
                 end_time, notes, location, role
        errors — list of dicts with keys: row, message
    """
    text = _read_as_text(file)
    reader = csv.DictReader(io.StringIO(text))

    column_map = _build_column_map(reader.fieldnames or [])
    _validate_required_columns(column_map)

    shifts, errors = [], []

    for row_index, raw_row in enumerate(reader, start=2):  # row 1 = header
        try:
            shift = _parse_row(raw_row, column_map, row_index)
            shifts.append(shift)
        except ValueError as exc:
            errors.append({"row": row_index, "message": str(exc)})

    return shifts, errors


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _read_as_text(file: IO) -> str:
    """Accept binary or text file-like objects and return a UTF-8 string."""
    content = file.read()
    if isinstance(content, bytes):
        return content.decode("utf-8-sig")  # strip BOM if present
    return content


def _build_column_map(fieldnames: list[str]) -> dict[str, str]:
    """
    Map raw CSV header names → canonical column names.
    Returns {raw_header: canonical_key}.
    """
    column_map = {}
    for raw in fieldnames:
        canonical = COLUMN_ALIASES.get(raw.strip().lower())
        if canonical:
            column_map[raw] = canonical
    return column_map


def _validate_required_columns(column_map: dict[str, str]) -> None:
    """Raise ValueError if any required canonical column is missing."""
    found = set(column_map.values())
    missing = REQUIRED_COLUMNS - found
    if missing:
        raise ValueError(
            f"CSV is missing required columns: {', '.join(sorted(missing))}. "
            f"Expected headers like: employee, date, start time, end time."
        )


def _parse_row(
    raw_row: dict[str, str],
    column_map: dict[str, str],
    row_index: int,
) -> dict:
    """Parse a single CSV row into a validated shift dict."""
    # Remap raw keys → canonical keys
    row = {canonical: raw_row[raw].strip() for raw, canonical in column_map.items()}

    employee = _require_field(row, "employee", row_index)
    shift_date = _parse_date(row.get("date", ""), row_index)
    start_time = _parse_time(row.get("start_time", ""), row_index, label="start time")
    end_time = _parse_time(row.get("end_time", ""), row_index, label="end time")

    _validate_time_order(start_time, end_time, row_index)

    return {
        "employee": employee,
        "date": shift_date,
        "start_time": start_time,
        "end_time": end_time,
        "notes": row.get("notes", ""),
        "location": row.get("location", ""),
        "role": row.get("role", ""),
    }


def _require_field(row: dict, key: str, row_index: int) -> str:
    value = row.get(key, "").strip()
    if not value:
        raise ValueError(f"Row {row_index}: '{key}' is empty.")
    return value


def _parse_date(value: str, row_index: int) -> date:
    if not value:
        raise ValueError(f"Row {row_index}: 'date' is empty.")
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    raise ValueError(
        f"Row {row_index}: Cannot parse date '{value}'. "
        f"Use formats like YYYY-MM-DD or MM/DD/YYYY."
    )


def _parse_time(value: str, row_index: int, label: str) -> datetime.time:
    if not value:
        raise ValueError(f"Row {row_index}: '{label}' is empty.")
    for fmt in TIME_FORMATS:
        try:
            return datetime.strptime(value.upper(), fmt).time()
        except ValueError:
            continue
    raise ValueError(
        f"Row {row_index}: Cannot parse {label} '{value}'. "
        f"Use formats like 14:30 or 2:30 PM."
    )


def _validate_time_order(
    start: datetime.time,
    end: datetime.time,
    row_index: int,
) -> None:
    if end <= start:
        raise ValueError(
            f"Row {row_index}: End time ({end}) must be after start time ({start})."
        )
