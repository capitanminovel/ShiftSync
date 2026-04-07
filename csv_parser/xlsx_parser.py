from __future__ import annotations
"""
csv_parser/xlsx_parser.py

Parses the columnar weekly schedule format used in the uploaded .xlsx files.

Layout:
  Row 1:  day-of-month numbers (e.g. 5, 6, 7 … 11), one per every 2 columns
  Row 2:  "SHIFT", "EMPLOYEE", "SHIFT", "EMPLOYEE" … (repeated headers)
  Row 3+: shift string (e.g. "9:30am - 4pm"), employee name pairs

The year and month are inferred from the filename (e.g. "4_5_26-4_11_26_Schedule.xlsx").
If they can't be parsed from the filename, the caller can pass them explicitly.

Usage:
    from csv_parser.xlsx_parser import parse_xlsx_schedule
    shifts, errors = parse_xlsx_schedule(filepath)
"""

import re
import os
from datetime import datetime, date
from typing import IO

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def parse_xlsx_schedule(
    filepath: str,
    year: int | None = None,
    month: int | None = None,
) -> tuple[list[dict], list[dict]]:
    """
    Parse a columnar weekly schedule .xlsx file.

    Args:
        filepath:  Path to the .xlsx file on disk.
        year:      Override year (inferred from filename if None).
        month:     Override month (inferred from filename if None).

    Returns:
        (shifts, errors)
        shifts — list of dicts: employee, date, start_time, end_time, notes
        errors — list of dicts: row, col_day, message
    """
    if year is None or month is None:
        inferred_year, inferred_month = _infer_year_month(filepath)
        year = year or inferred_year
        month = month or inferred_month

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 3:
        return [], [{"row": 1, "col_day": None, "message": "File has fewer than 3 rows — nothing to parse."}]

    day_numbers = _extract_day_numbers(rows[0])
    data_rows = rows[2:]  # skip day-number row and header row

    shifts, errors = [], []
    num_days = len(day_numbers)

    for row_idx, row in enumerate(data_rows, start=3):  # 1-based; row 1 = day row
        for day_col, day_num in enumerate(day_numbers):
            shift_idx = day_col * 2
            emp_idx = shift_idx + 1

            shift_str = _cell(row, shift_idx)
            employee = _cell(row, emp_idx)

            if not employee or not shift_str:
                continue
            if _is_metadata(employee) or _is_metadata(shift_str):
                continue

            try:
                shift_date = date(year, month, int(day_num))
            except ValueError as exc:
                errors.append({
                    "row": row_idx,
                    "col_day": day_num,
                    "message": f"Invalid date: day {day_num} in {month}/{year} — {exc}",
                })
                continue

            try:
                start_time, end_time = _parse_shift_string(shift_str)
            except ValueError as exc:
                errors.append({
                    "row": row_idx,
                    "col_day": day_num,
                    "message": f"Could not parse shift '{shift_str}' for {employee}: {exc}",
                })
                continue

            shifts.append({
                "employee":   employee.strip(),
                "date":       shift_date,
                "start_time": start_time,
                "end_time":   end_time,
                "notes":      str(shift_str).strip(),
                "location":   "",
                "role":       "",
            })

    return shifts, errors


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _cell(row: tuple, idx: int) -> str | None:
    """Safe indexed access into a row tuple; returns stripped string or None."""
    if idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    val = str(val).strip()
    return val if val else None


def _extract_day_numbers(day_row: tuple) -> list[int]:
    """
    Extract integer day-of-month values from row 1.
    Only even-indexed cells (0, 2, 4 …) contain day numbers; odd cells are None.
    """
    numbers = []
    for i in range(0, len(day_row), 2):
        val = day_row[i]
        try:
            numbers.append(int(val))
        except (TypeError, ValueError):
            pass
    return numbers


def _is_metadata(value: str) -> bool:
    """Return True for non-shift annotation strings like 'UNAVAILABLE:'."""
    keywords = ("unavailable", "note", "off", "vacation", "pto")
    return any(value.lower().startswith(k) for k in keywords)


def _infer_year_month(filepath: str) -> tuple[int, int]:
    """
    Try to infer month and year from filenames like:
      4_5_26-4_11_26_Schedule.xlsx
      2026-04-05_week.xlsx
    Returns (year, month). Falls back to today if parsing fails.
    """
    name = os.path.basename(filepath)

    # Pattern: M_D_YY or MM_DD_YYYY at the start
    m = re.search(r'(\d{1,2})[_\-](\d{1,2})[_\-](\d{2,4})', name)
    if m:
        month_raw, _, year_raw = int(m.group(1)), int(m.group(2)), int(m.group(3))
        year = year_raw + 2000 if year_raw < 100 else year_raw
        return year, month_raw

    # ISO pattern: YYYY-MM-DD
    m = re.search(r'(\d{4})[_\-](\d{2})[_\-](\d{2})', name)
    if m:
        return int(m.group(1)), int(m.group(2))

    today = datetime.today()
    return today.year, today.month


# ---------------------------------------------------------------------------
# Time / shift string parsing
# ---------------------------------------------------------------------------

def _parse_shift_string(shift_str: str) -> tuple[datetime.time, datetime.time]:
    """
    Parse strings like:
        "9:30am - 4pm"
        "4:30/5pm - 8+pm"
        "12pm - 8+pm"

    Returns (start_time, end_time) as datetime.time objects.
    Raises ValueError if parsing fails.
    """
    parts = re.split(r'\s*-\s*', shift_str.strip(), maxsplit=1)
    if len(parts) != 2:
        raise ValueError(f"Expected 'start - end' format, got: '{shift_str}'")
    start = _parse_time_token(parts[0])
    end   = _parse_time_token(parts[1])
    if start is None:
        raise ValueError(f"Cannot parse start time: '{parts[0]}'")
    if end is None:
        raise ValueError(f"Cannot parse end time: '{parts[1]}'")
    return start, end


def _parse_time_token(tok: str) -> datetime.time | None:
    """
    Parse a single time token:
        "9:30am"  → 09:30
        "4pm"     → 16:00
        "8+pm"    → 20:00   ('+' means 'or later'; we drop it)
        "4:30/5pm"→ 16:30   (slash variant; we take the earlier value)
    """
    tok = tok.strip().lower()
    tok = tok.replace('+', '')            # "8+pm" → "8pm"

    if '/' in tok:
        # "4:30/5pm" → keep first part, reattach am/pm suffix
        suffix_m = re.search(r'[ap]m$', tok)
        suffix = suffix_m.group() if suffix_m else ''
        tok = tok.split('/')[0] + suffix  # "4:30pm"

    tok = tok.upper()
    for fmt in ('%I:%M%p', '%I%p'):
        try:
            return datetime.strptime(tok, fmt).time()
        except ValueError:
            continue
    return None
